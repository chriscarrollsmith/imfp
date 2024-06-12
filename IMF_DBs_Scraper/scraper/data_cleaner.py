import re
import logging
import os
import json
import shutil
import time
import zipfile
import gzip
import pandas as pd
from zipfile import BadZipFile
from pandas.errors import EmptyDataError
import win32com.client as win32
from datetime import datetime
from pathlib import Path
import xlrd
import xlwt
import xlsxwriter
from openpyxl import load_workbook, Workbook
from typing import List
import requests


class DataCleaner:

    def __init__(self):
        pass

    @staticmethod
    def get_internet_banking_values_from_api(api_url: str) -> pd.DataFrame:
        # Send request to API and get response
        response = requests.get(api_url)

        # Check if response status code is 200 (OK)
        if response.status_code != 200:
            print(f"Request failed with status code {response.status_code}")
            return None

        # Parse response JSON to extract required data
        data = {"category": [], "dataYear": [], "shortName": [], "isoCode": [], "value": []}
        for record in response.json():
            data["category"].append("Internet Banking")
            data["dataYear"].append(record["dataYear"])
            data["shortName"].append(record["shortName"])
            data["isoCode"].append(record["isoCode"])
            data["value"].append(record["answer"][0]["value"])

        # Convert data to pandas DataFrame
        df = pd.DataFrame(data)

        return df

    @staticmethod
    def get_iso_codes_from_json(file_path):
        print(f"Opening JSON file at {file_path}...")

        # Open the JSON file and load its contents as a Python object
        with open(file_path, 'r') as f:
            json_data = json.load(f)

        print("Extracting 'IsoCode' values...")

        # Extract the "IsoCode" values from each object in the list
        iso_codes = [country['IsoCode'] for country in json_data]

        print(f"Found {len(iso_codes)} 'IsoCode' values")

        # Return the list of "IsoCode" values
        return iso_codes

    @staticmethod
    def save_api_response_to_json(api_url, output_path, output_file_name):
        # Make a request to the API endpoint
        response = requests.get(api_url)

        # Parse the response as JSON
        response_json = response.json()

        # Combine the output path and file name
        output = os.path.join(output_path, output_file_name)

        # Open the output file and write the JSON data
        with open(output, 'w') as outfile:
            json.dump(response_json, outfile, indent=4)

        # Print a success message
        print(f"JSON data saved to {output}")

    @staticmethod
    def combine_excel_files_into_one(folder_path):
        print(f"Combining Excel files in folder: {folder_path}")

        # Get a list of all Excel files in the folder
        excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

        # Initialize an empty DataFrame to hold the combined data
        combined_data = pd.DataFrame()

        # Loop over each Excel file and append its data to the combined DataFrame
        for file in excel_files:
            file_path = os.path.join(folder_path, file)
            print(f"Reading Excel file: {file_path}")
            data = pd.read_excel(file_path)
            combined_data = pd.concat([combined_data, data], ignore_index=True)

        # Save the combined data to a new Excel file with the same name as the folder
        combined_file_path = os.path.join(folder_path, f'{os.path.basename(folder_path)}.xlsx')
        print(f"Saving combined data to Excel file: {combined_file_path}")
        try:
            combined_data.to_excel(combined_file_path, index=False)
            print('Saved Excel file successfully!')
        except ValueError as e:
            print(f'ValueError:{e}')
            print('will try to save combined data to CSV file instead')
            combined_file_path = os.path.join(folder_path, f'{os.path.basename(folder_path)}.csv')
            combined_data.to_csv(combined_file_path, index=False)
            print('Saved CSV file successfully!')

        # Delete the original Excel files
        for file in excel_files:
            file_path = os.path.join(folder_path, file)
            print(f"Deleting original Excel file: {file_path}")
            os.remove(file_path)

        print(f"Done the combining procedure in {folder_path}.")

    @staticmethod
    def delete_last_created_folder(path):
        # Change the current working directory to the target path
        os.chdir(path)

        # Create an empty list to store creation times of each directory
        l = []

        # Loop through all directories in the target path
        for dir in os.listdir(path):
            # Get the absolute path of the directory
            abs_path = os.path.abspath(dir)
            # Add the creation time of the directory to the list
            l.append(os.path.getctime(abs_path))

        # Find the creation time of the last created directory
        dir_to_delete = max(l)

        # Loop through all directories in the target path again
        for dir in os.listdir(path):
            # Get the absolute path of the directory
            abs_path = os.path.abspath(dir)
            # If the directory was the last created, delete it and break the loop
            if os.path.getctime(abs_path) == dir_to_delete:
                print(f"Deleting last created folder {abs_path}")
                shutil.rmtree(abs_path)
                break

    @staticmethod
    # this function needs work (do not use it)
    def split_excel_into_files_and_group_by_col(input_file, output_folder, group_by_col):
        print("Reading input Excel file...")
        df = pd.read_excel(input_file, sheet_name=0)

        print(f"Grouping data by {group_by_col}...")
        grouped = df.groupby(f'{group_by_col}')

        print("Writing groups to separate Excel files...")
        for name, group in grouped:
            output_file = f"{output_folder}/{name}/{name}.xlsx"
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            group.to_excel(output_file, index=False)

        print("Reading metadata from the second sheet...")
        metadata_df = pd.read_excel(input_file, sheet_name=1)

        print("Creating JSON files for each metadata entry...")
        for _, row in metadata_df.iterrows():
            row = row.fillna("")  # Fill NaN values with empty strings
            code = row["Code"]
            metadata = {
                "url": "https://databank.worldbank.org/id/2ddc971b?Code=SG.OPN.BANK.EQ&report_name=Gender_Indicators_Report&populartype=series#",
                "indicator": code,
                "additional_data_sources": [],
                "table": row["Indicator Name"],
                "description": row["Long definition"],
                "tags": [{"name": row["Topic"].lower() if row["Topic"] else "unclassified"}],
                "limitations": row["Limitations and exceptions"],
                "concept": row["Statistical concept and methodology"],
                "periodicity": row["Periodicity"],
                "topic": row["Topic"],
                "created": "",
                "last_modified": ""
            }

            # Save the metadata JSON file in the appropriate folder
            metadata_file = f"{output_folder}/{code}/Metadata_{code}.json"
            print(f"Making folder based on 'Code'")
            os.makedirs(os.path.dirname(metadata_file), exist_ok=True)
            with open(metadata_file, 'w') as f:
                json.dump(metadata, f, indent=4)
            print(f"Saved metadata for {code}.")

    @staticmethod
    def get_excel_csv_files_paths(master_dir):
        # Initialize an empty list to store file paths
        file_paths = []

        # Loop through all the files and subdirectories in the master directory
        for root, dirs, files in os.walk(master_dir):
            for file in files:
                # Check if the file is an .xlsx or .csv file
                if file.endswith('.xlsx') or file.endswith('.csv'):
                    # If so, append the file path to the list
                    file_paths.append(os.path.join(root, file))

        # Return the list of file paths
        return file_paths

    @staticmethod
    def create_sub_folder(working_dir, folder_name):
        # Combine the working directory and subfolder name to get the full path
        sub_working_dir = os.path.join(working_dir, folder_name)

        # If the subfolder doesn't exist, create it
        if not os.path.exists(sub_working_dir):
            print(f"Creating subfolder {sub_working_dir}")
            os.makedirs(sub_working_dir)

        # Return the full path of the subfolder
        return sub_working_dir

    @staticmethod
    def create_subfolders(directory_path: str, subfolders: List[str]) -> str:
        # Start with the given directory path
        current_path = directory_path

        # Loop through each subfolder in the list
        for subfolder in subfolders:
            # Append the subfolder to the current path
            current_path = os.path.join(current_path, subfolder)

            # If the subfolder doesn't exist, create it
            if not os.path.exists(current_path):
                print(f"Creating subfolder {current_path}")
                os.mkdir(current_path)

        # Return the final path with all subfolders
        return current_path

    @staticmethod
    def update_dates_in_metadata_files(master_dir):
        def reformat_date(date_str):
            """
            Helper function to reformat a date string from dd-mm-yyyy to yyyy-mm-dd.
            """
            date_obj = datetime.strptime(date_str, '%d-%m-%Y')
            return date_obj.strftime('%Y-%m-%d')

        def update_dates_in_json_file(file_path):
            """
            Helper function to update the 'created' and 'last_modified' dates in a JSON file.
            """
            # Load the JSON data from the file
            with open(file_path, 'r') as json_file:
                data = json.load(json_file)

            # If 'created' field exists, reformat the date and update it in the JSON data
            if 'created' in data:
                data['created'] = reformat_date(data['created'])

            # If 'last_modified' field exists, reformat the date and update it in the JSON data
            if 'last_modified' in data:
                data['last_modified'] = reformat_date(data['last_modified'])

            # Write the updated JSON data back to the file
            with open(file_path, 'w') as json_file:
                json.dump(data, json_file, indent=4)

        # Loop through all folders in the master folder
        for foldername in os.listdir(master_dir):
            folder_path = os.path.join(master_dir, foldername)
            # If the item is a directory
            if os.path.isdir(folder_path):
                # Loop through all files in the directory
                for filename in os.listdir(folder_path):
                    # If the file is a metadata JSON file
                    if filename.startswith("Metadata_") and filename.endswith(".json"):
                        file_path = os.path.join(folder_path, filename)
                        # Update the dates in the metadata JSON file
                        print(f"Updating dates in {file_path}")
                        update_dates_in_json_file(file_path)

    @staticmethod
    def convert_tsv_to_csv(tsv_file_path):
        ext = '.tsv'
        os.chdir(tsv_file_path)
        success = True

        for file in os.listdir(tsv_file_path):
            if file.endswith(ext):
                try:
                    # read the .tsv file into a dataframe
                    df = pd.read_csv(file, sep="\t")
                    # write the dataframe to a .csv file
                    df.to_csv(f"{file}.csv", index=False)
                    print(f'{file} converted from .tsv to .csv successfully!')
                except Exception as e:
                    success = False
                    print(f"Error converting {file}: {e}")

        return success

    @staticmethod
    def convert_csv_to_xlsx(csv_file, output_path, xlsx_file_name):
        try:
            # reading the csv file
            csv_df = pd.read_csv(csv_file)
            # creating an output excel file
            result_xlsx = pd.ExcelWriter(os.path.join(output_path, xlsx_file_name))
            # converting the csv file to an excel file
            csv_df.to_excel(result_xlsx, index=False)
            # saving the excel file
            result_xlsx.save()
            print(f'{csv_file} converted to {xlsx_file_name} successfully!')
            return True
        except Exception as e:
            print(f"Error converting {csv_file}: {e}")
            return False

    @staticmethod
    def convert_txt_to_csv(txt_file_path):
        ext = '.txt'
        os.chdir(txt_file_path)
        success = True

        for file in os.listdir(txt_file_path):
            if file.endswith(ext):
                try:
                    df = pd.read_csv(file, header=None, delim_whitespace=True, low_memory=False)
                    df.to_csv(file + '.csv', index=False, header=None)
                    print(f'{file} changed from .txt to .csv successfully')
                except Exception as e:
                    success = False
                    print(f"Error converting {file}: {e}")

        return success

    @staticmethod
    def convert_txt_to_xlsx(txt_file_path):
        ext = '.txt'
        os.chdir(txt_file_path)
        success = True

        for file in os.listdir(txt_file_path):
            if file.endswith(ext):
                try:
                    df = pd.read_csv(file, header=None, delim_whitespace=True, low_memory=False)
                    df.to_excel(file + '.xlsx', index=False, header=None)
                    print('file changed from .txt to .xlsx successfully')
                except Exception as e:
                    success = False
                    print(f"Error converting {file}: {e}")

        return success

    @staticmethod
    def convert_xls_to_xlsx_win32(master_dir):
        success = True

        for folder in os.listdir(master_dir):
            current_dir = os.path.join(master_dir, folder)

            if os.path.isdir(current_dir):
                for file in os.listdir(current_dir):
                    if file.endswith('.xls'):
                        try:
                            fname = os.path.join(current_dir, file)
                            excel = win32.gencache.EnsureDispatch('Excel.Application')
                            wb = excel.Workbooks.Open(fname)

                            wb.SaveAs(fname + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
                            wb.Close()  # FileFormat = 56 is for .xls extension
                            excel.Application.Quit()
                            print(f'{file} converted to {file[:-4]}.xlsx successfully!')

                        except Exception as e:
                            success = False
                            print(f"Error converting {file}: {e}")

        return success

    @staticmethod
    def move_folders_without_data_files(master_folder, destination_folder):
        # Allowed file extensions
        allowed_extensions = {'.xlsx', '.xls', '.xlsm', '.csv'}

        # Create the destination folder if it doesn't exist
        if not os.path.exists(destination_folder):
            os.makedirs(destination_folder)
            print(f"Created destination folder: '{destination_folder}'")

        # Iterate over subfolders in the master folder
        for root, dirs, files in os.walk(master_folder):
            for directory in dirs:
                dir_path = os.path.join(root, directory)
                should_move = True

                # Iterate over files in the current subfolder
                for file in os.listdir(dir_path):
                    # Get file extension and convert to lowercase
                    file_extension = os.path.splitext(file)[1].lower()

                    # Check if file extension is in the allowed extensions list
                    if file_extension in allowed_extensions:
                        should_move = False
                        break

                # Move the folder if it doesn't contain any files with allowed extensions
                if should_move:
                    source = dir_path
                    destination = os.path.join(destination_folder, directory)
                    print(f"Moving '{source}' to '{destination}' because folder has no data")
                    shutil.move(source, destination)

    @staticmethod
    def move_zero_table_rows_folders(source_dir, destination_dir):
        source_dir = Path(source_dir)
        destination_dir = Path(destination_dir)
        allowed_extensions = {'.xlsx', '.xls', '.xlsm', '.csv', '.XLSX', '.XLS', '.XLSM', '.CSV'}

        # Iterate through all subfolders in the source directory
        for subfolder in source_dir.iterdir():
            if subfolder.is_dir():
                # Get data files with allowed extensions
                data_files = [f for f in subfolder.glob("*") if f.suffix in allowed_extensions]

                # Initialize a flag to check if all data files have zero rows
                all_files_have_zero_rows = True

                # Iterate through data files
                for data_file in data_files:
                    try:
                        # Read the file using pandas
                        if data_file.suffix.lower() in ['.xlsx', '.xls', '.xlsm']:
                            df = pd.read_excel(data_file)
                        else:
                            df = pd.read_csv(data_file)

                        # Check if the data file has non-zero table rows (excluding the header)
                        if df.shape[0] != 0:
                            all_files_have_zero_rows = False
                            break
                    except Exception as e:
                        print(f"Error processing file {data_file}: {e}")
                        all_files_have_zero_rows = False
                        break

                # Move the subfolder if all data files have zero rows
                if all_files_have_zero_rows:
                    destination_subfolder = destination_dir / subfolder.name
                    print(
                        f"Moving {subfolder} to {destination_subfolder} because one of its data file contains zero table rows")
                    shutil.move(str(subfolder), str(destination_dir))

    @staticmethod
    def list_folder_structure(path):
        for root, dirs, files in os.walk(path):
            level = root.replace(path, '').count(os.sep)
            indent = ' ' * 4 * (level)
            print('{}{}/'.format(indent, os.path.basename(root)))
            subindent = ' ' * 4 * (level + 1)
            for f in files:
                print('{}{}'.format(subindent, f))

    @staticmethod
    def rename_metadata_files(master_dir, old_starts_with, new_starts_with):
        for root, dirs, files in os.walk(master_dir):
            # iterate over all subdirectories in master_dir
            for dir_name in dirs:
                sub_dir = os.path.join(root, dir_name)
                # iterate over all files in subdirectory
                for file_name in os.listdir(sub_dir):
                    if file_name.startswith(old_starts_with) and file_name.endswith('.json'):
                        # create new file name by replacing "Metadata_" with "MetaORIdata_"
                        new_file_name = file_name.replace(old_starts_with, new_starts_with)
                        # rename file
                        os.rename(os.path.join(sub_dir, file_name), os.path.join(sub_dir, new_file_name))
                        print(f"Renamed {file_name} to {new_file_name} in {sub_dir}")

    @staticmethod
    def move_data_files(download_dir, target_dir):
        # Check if download_dir exists
        if not os.path.exists(download_dir):
            print(f"The download directory '{download_dir}' does not exist.")
            logging.error(f"The download directory '{download_dir}' does not exist.")
            return

        # Check if target_dir exists, create it if not
        if not os.path.exists(target_dir):
            os.makedirs(target_dir)
            print(f"The target directory '{target_dir}' was created.")
            logging.info(f"The target directory '{target_dir}' was created.")

        # Move files with the given extension from download_dir to target_dir
        moved_files = 0
        for filename in os.listdir(download_dir):
            if filename.lower().endswith(('.csv', '.xlsx', '.xls', '.xlsm', '.tsv', '.zip')):
                src_file = os.path.join(download_dir, filename)
                dest_file = os.path.join(target_dir, filename)
                shutil.move(src_file, dest_file)
                print(f"File '{filename}' was moved from '{download_dir}' to '{target_dir}'.")
                logging.info(f"File '{filename}' was moved from '{download_dir}' to '{target_dir}'.")
                moved_files += 1

        # Print and log the number of moved files
        if moved_files == 0:
            print(f"No files with the allowed extensions were found in '{download_dir}'.")
            logging.warning(f"No files with the allowed extensions were found in '{download_dir}'.")
        else:
            print(
                f"{moved_files} files with the allowed extensions were moved from '{download_dir}' to '{target_dir}'.")
            logging.info(
                f"{moved_files} files with the allowed extensions were moved from '{download_dir}' to '{target_dir}'.")

    @staticmethod
    def replace_subscript_numbers(dir_path):
        """
        This function iterates over the subdirectories of the given directory
        and replaces any subscript numbers in their names with their corresponding
        regular numbers. The new names are saved to disk.
        """

        # Iterate over the subdirectories of the given directory
        for subdir in os.listdir(dir_path):
            # Check if the subdirectory is a directory (not a file)
            if os.path.isdir(os.path.join(dir_path, subdir)):
                # Generate the new name for the subdirectory by replacing subscript numbers with regular numbers
                new_name = re.sub("₀", "0", subdir)
                new_name = re.sub("₁", "1", new_name)
                new_name = re.sub("₂", "2", new_name)
                new_name = re.sub("₃", "3", new_name)
                new_name = re.sub("₄", "4", new_name)
                new_name = re.sub("₅", "5", new_name)
                new_name = re.sub("₆", "6", new_name)
                new_name = re.sub("₇", "7", new_name)
                new_name = re.sub("₈", "8", new_name)
                new_name = re.sub("₉", "9", new_name)

                # Print the old and new names for debugging purposes
                print(f"Renaming '{subdir}' to '{new_name}'...")

                # Rename the subdirectory on disk
                os.rename(os.path.join(dir_path, subdir), os.path.join(dir_path, new_name))

        # Print a message indicating that the function has finished running
        print("Done!")

    @staticmethod
    def move_subdirs_with_more_than_two_files(master_dir, target_dir):
        """
        Move all subdirectories in master_dir that contain more than 2 files to target_dir.
        """
        for subdir, _, files in os.walk(master_dir):
            if len(files) > 2:
                # Move the subdirectory to the target directory
                target_subdir = os.path.join(target_dir, os.path.basename(subdir))
                shutil.move(subdir, target_subdir)

    @staticmethod
    def move_large_files(origin_dir, target_dir):
        max_size = 100 * 1024 * 1024
        if not os.path.exists(target_dir):
            os.makedirs(target_dir)

        for filename in os.listdir(origin_dir):
            if filename.endswith(".json"):
                origin_path = os.path.join(origin_dir, filename)
                if os.path.getsize(origin_path) >= max_size:
                    target_path = os.path.join(target_dir, filename)
                    shutil.move(origin_path, target_path)
                    print(f'moved {origin_path} to {target_path}')
                    print('')

    @staticmethod
    def generate_file_name(table_name):
        # Remove illegal characters and truncate the string to a maximum of 70 characters
        sanitized_name = re.sub(r"[\\/:*?\"<>|'!@#$%^&()_+=\.\s,;\[\]{}—-]+", "_", table_name)[:70]

        # Replace subscript numbers with regular numbers
        sanitized_name = re.sub("₀", "0", sanitized_name)
        sanitized_name = re.sub("₁", "1", sanitized_name)
        sanitized_name = re.sub("₂", "2", sanitized_name)
        sanitized_name = re.sub("₃", "3", sanitized_name)
        sanitized_name = re.sub("₄", "4", sanitized_name)
        sanitized_name = re.sub("₅", "5", sanitized_name)
        sanitized_name = re.sub("₆", "6", sanitized_name)
        sanitized_name = re.sub("₇", "7", sanitized_name)
        sanitized_name = re.sub("₈", "8", sanitized_name)
        sanitized_name = re.sub("₉", "9", sanitized_name)

        # Remove repeated underscores
        sanitized_name = re.sub(r"_+", "_", sanitized_name)

        # Remove leading and trailing underscores if present
        file_name = sanitized_name.strip("_")

        return file_name

    @staticmethod
    def move_folders_without_metadata_files(master_folder, destination_folder):
        # Allowed file extensions
        allowed_extensions = {'.json'}

        # Create the destination folder if it doesn't exist
        if not os.path.exists(destination_folder):
            os.makedirs(destination_folder)
            print(f"Created destination folder: '{destination_folder}'")

        # Iterate over subfolders in the master folder
        for root, dirs, files in os.walk(master_folder):
            for directory in dirs:
                dir_path = os.path.join(root, directory)
                should_move = True

                # Iterate over files in the current subfolder
                for file in os.listdir(dir_path):
                    # Get file extension and convert to lowercase
                    file_extension = os.path.splitext(file)[1].lower()

                    # Check if file extension is in the allowed extensions list
                    if file_extension in allowed_extensions:
                        should_move = False
                        break

                # Move the folder if it doesn't contain any files with allowed extensions
                if should_move:
                    source = dir_path
                    destination = os.path.join(destination_folder, directory)
                    print(f"Moving '{source}' to '{destination}' because folder has no metadata")
                    shutil.move(source, destination)

    @staticmethod
    def rename_item(old_path, pattern=r"[-_—]+", replace_with="_"):
        new_name = re.sub(pattern, replace_with, os.path.basename(old_path))
        if new_name != os.path.basename(old_path):
            new_path = os.path.join(os.path.dirname(old_path), new_name)
            os.rename(old_path, new_path)
            print(f'Renamed: {old_path} -> {new_path}')
            return new_path
        return old_path

    @staticmethod
    def rename_files_and_folders(master_dir):
        # Rename the master directory itself
        master_dir = DataCleaner.rename_item(master_dir, r"[-_—]+")

        # Rename the contents of the master directory
        for root, dirs, files in os.walk(master_dir):
            for file in files:
                old_file_path = os.path.join(root, file)
                DataCleaner.rename_item(old_file_path, r"[-_—]+")

            for folder in list(dirs):
                old_folder_path = os.path.join(root, folder)
                new_folder_path = DataCleaner.rename_item(old_folder_path, r"[-_—]+")
                if new_folder_path != old_folder_path:
                    dirs.remove(folder)
                    dirs.append(os.path.basename(new_folder_path))

    @staticmethod
    def get_folders_numbers(source_dir):
        source_dir = Path(source_dir)
        folder_numbers = []

        # Iterate through all subfolders in the source directory
        for subfolder in source_dir.iterdir():
            if subfolder.is_dir():
                # Extract the number at the end of the folder name using regex
                match = re.search(r'_(\d+)$', subfolder.name)
                if match:
                    folder_number = int(match.group(1))
                    folder_numbers.append(folder_number)

        # Sort the folder_numbers list in ascending order
        folder_numbers.sort()

        print(folder_numbers)
        print(f'length: {len(folder_numbers)}')
        return folder_numbers

    @staticmethod
    def move_data_files(download_dir, target_dir):
        # Check if download_dir exists
        if not os.path.exists(download_dir):
            logging.error(f"The download directory '{download_dir}' does not exist.")
            return

        # Check if target_dir exists, create it if not
        if not os.path.exists(target_dir):
            os.makedirs(target_dir)
            logging.info(f"The target directory '{target_dir}' was created.")

        # Move files with the given extension from download_dir to target_dir
        moved_files = 0
        for filename in os.listdir(download_dir):
            if filename.lower().endswith(('.csv', '.xlsx', '.xls', '.xlsm', '.zip')):
                src_file = os.path.join(download_dir, filename)
                dest_file = os.path.join(target_dir, filename)
                shutil.move(src_file, dest_file)
                logging.info(f"File '{filename}' was moved from '{download_dir}' to '{target_dir}'.")
                moved_files += 1

        # Print and log the number of moved files
        if moved_files == 0:
            logging.warning(f"No files with the allowed extensions were found in '{download_dir}'.")
        else:
            logging.info(
                f"{moved_files} files with the allowed extensions were moved from '{download_dir}' to '{target_dir}'.")

    @staticmethod
    def move_metadata_files(download_dir, target_dir):
        # Check if download_dir exists
        if not os.path.exists(download_dir):
            logging.error(f"The download directory '{download_dir}' does not exist.")
            return

        # Check if target_dir exists, create it if not
        if not os.path.exists(target_dir):
            os.makedirs(target_dir)
            logging.info(f"The target directory '{target_dir}' was created.")

        # Move files with the given extension from download_dir to target_dir
        moved_files = 0
        for filename in os.listdir(download_dir):
            if filename.lower().endswith(('.json')):
                src_file = os.path.join(download_dir, filename)
                dest_file = os.path.join(target_dir, filename)
                shutil.move(src_file, dest_file)
                logging.info(f"File '{filename}' was moved from '{download_dir}' to '{target_dir}'.")
                moved_files += 1

        # Print and log the number of moved files
        if moved_files == 0:
            logging.warning(f"No files with the allowed extensions were found in '{download_dir}'.")
        else:
            logging.info(
                f"{moved_files} files with the allowed extensions were moved from '{download_dir}' to '{target_dir}'.")

    @staticmethod
    def move_large_files(origin_dir, target_dir):
        max_size = 100 * 1024 * 1024
        if not os.path.exists(target_dir):
            os.makedirs(target_dir)

        for filename in os.listdir(origin_dir):
            if filename.endswith(".json"):
                origin_path = os.path.join(origin_dir, filename)
                if os.path.getsize(origin_path) >= max_size:
                    target_path = os.path.join(target_dir, filename)
                    shutil.move(origin_path, target_path)
                    print(f'moved {origin_path} to {target_path}')
                    print('')

    @staticmethod
    def extract_here(master_dir, extension):
        # change path from working dir to dir with files
        os.chdir(master_dir)
        all_subdirs = []
        # loop through items in dir
        for items in os.walk(master_dir):
            # check for specified extension
            all_subdirs.append(items[0])
        for subdir in all_subdirs:
            os.chdir(subdir)
            for item in os.listdir(subdir):
                try:
                    if item.endswith(extension) and extension == '.zip':
                        print('to extract: ' + item)
                        # get full path of file
                        file_name = os.path.abspath(item)
                        # create zipfile object
                        zip_ref = zipfile.ZipFile(file_name)
                        # extract file to current dir
                        zip_ref.extractall(subdir)
                        print('extracted ' + item + ' file')
                        # close file
                        zip_ref.close()
                        # delete zipped file
                        os.remove(file_name)
                        print('deleted ' + item + ' file')
                        # revert to master dir
                        os.chdir(master_dir)
                    elif item.endswith(extension) and extension == '.gz':
                        print('to extract: ' + item)
                        # get file name
                        file_name = item
                        # extract file to current dir
                        with gzip.open(file_name, 'r') as f_in:
                            with open(os.path.join(subdir, file_name[:-3] + '.txt'), 'wb') as f_out:
                                shutil.copyfileobj(f_in, f_out)
                        print('extracted ' + item + ' file')
                        # delete zipped file
                        os.remove(file_name)
                        print('deleted ' + item + ' file')
                        # revert to master dir
                        os.chdir(master_dir)
                except BadZipFile:
                    print(file_name + ' is corrupted. You will have to extract it manually.')
                    continue

    @staticmethod
    def unzip_files(master_dir, extension):
        # change path from working dir to dir with files
        os.chdir(master_dir)
        all_subdirs = []
        # loop through items in dir
        for items in os.walk(master_dir):
            # check for ".zip" extension
            all_subdirs.append(items[0])
        for subdir in all_subdirs:
            os.chdir(subdir)
            for item in os.listdir(subdir):
                try:
                    if item.endswith(extension) and extension == '.zip':
                        print('to unzip: ' + item)
                        # get full path of files
                        file_name = os.path.abspath(item)
                        # create zipfile object
                        zip_ref = zipfile.ZipFile(file_name)
                        # create folder with file_name
                        os.makedirs(item[:-4])
                        new_dir = os.path.join(subdir, item[:-4])
                        os.chdir(new_dir)
                        # extract file to dir
                        zip_ref.extractall(new_dir)
                        print('unzipped ' + item + extension + ' file')
                        # close file
                        zip_ref.close()
                        # delete zipped file
                        os.remove(file_name)
                        print('deleted ' + item + ' file')
                        # revert to master dir
                        os.chdir(master_dir)
                    elif item.endswith(extension) and extension == '.gz':
                        print('to unzip: ' + item)
                        # get file name
                        file_name = item
                        # create folder with file_name
                        os.makedirs(item[:-3])
                        new_dir = os.path.join(subdir, item[:-3])
                        # extract file to dir
                        with gzip.open(file_name, 'r') as f_in:
                            with open(os.path.join(new_dir, file_name + '.txt'), 'wb') as f_out:
                                shutil.copyfileobj(f_in, f_out)
                        print('unzipped ' + item + ' file')
                        # delete zipped file
                        os.remove(file_name)
                        print('deleted ' + item + ' file')
                        # revert to master dir
                        os.chdir(master_dir)
                except BadZipFile:
                    print(file_name + ' is corrupted u will have to unzip it manually.')
                    continue

    @staticmethod
    def delete_zero_table_row_folder(folder_path):
        folder_path = Path(folder_path)
        allowed_extensions = {'.xlsx', '.xls', '.xlsm', '.csv', '.XLSX', '.XLS', '.XLSM', '.CSV'}

        if folder_path.is_dir():
            # Get data files with allowed extensions
            data_files = [f for f in folder_path.glob("*") if f.suffix in allowed_extensions]

            # Initialize a flag to check if all data files have zero rows
            all_files_have_zero_rows = True

            # Iterate through data files
            for data_file in data_files:
                try:
                    # Read the file using pandas
                    if data_file.suffix.lower() in ['.xlsx', '.xls', '.xlsm']:
                        df = pd.read_excel(data_file)
                    else:
                        df = pd.read_csv(data_file)

                    # Check if the data file has non-zero table rows (excluding the header)
                    if df.shape[0] != 0:
                        all_files_have_zero_rows = False
                        break
                except Exception as e:
                    print(f"Error processing file {data_file}: {e}")
                    all_files_have_zero_rows = False
                    break

            # Delete the folder if all data files have zero rows
            if all_files_have_zero_rows:
                print(f"Deleting {folder_path} because its one/many data files contain zero table rows")
                shutil.rmtree(str(folder_path))
            else:
                print(f"Not deleting {folder_path} because not all data files contain zero table rows")

    # DATA MANIPULATION AND CLEANING
    # Step zero:
    @staticmethod
    def delete_specific_sheets(source_dir, output_dir, keywords_to_delete=None):
        if keywords_to_delete is None:
            keywords_to_delete = ['Note', 'Round', '%', 'type', 'Tabulation', 'Summary', 'Displacement', 'Analysis',
                                  'Content', 'Total', 'Bi-weekly', 'Monthly', 'Governorate', 'Dictionary', 'ILA',
                                  'Data Dictionary', 'Summaries', 'Pivot', 'Matrix', 'Admin2', 'new', 'Map',
                                  'Description', 'Sheet2', 'Metadata']
        # Check if output directory exists, and create it if it doesn't
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        def copy_worksheets(xlrd_book, xlwt_book, sheet_name):
            xlrd_sheet = xlrd_book.sheet_by_name(sheet_name)
            xlwt_sheet = xlwt_book.add_sheet(xlrd_sheet.name)

            for row in range(xlrd_sheet.nrows):
                for col in range(xlrd_sheet.ncols):
                    xlwt_sheet.write(row, col, xlrd_sheet.cell_value(row, col))

        source_path = Path(source_dir)
        output_path = Path(output_dir)
        regexes_to_delete = [re.compile(fr"{re.escape(keyword)}", re.IGNORECASE) for keyword in keywords_to_delete]

        for subdir in source_path.glob('**/*'):
            if subdir.is_dir():
                print(f"Processing folder to delete unwanted excel sheets: {subdir}")
                relative_subdir = subdir.relative_to(source_path)
                output_subdir = output_path / relative_subdir
                output_subdir.mkdir(parents=True, exist_ok=True)

                for file in subdir.glob('*.*'):
                    if file.name.startswith('~$'):
                        continue

                    file_ext = file.suffix.lower()
                    output_file = output_subdir / file.name
                    # Copy non-Excel files to the output director
                    if file_ext not in ['.xls', '.xlsx', '.xlsm']:
                        shutil.copy(file, output_file)
                        print(f"Copied file: {file} to {output_file}")

                    if file_ext in ['.xls']:
                        print(f"Processing .xls file for deletion of unwanted sheets: {file}")

                        try:
                            xlrd_book = xlrd.open_workbook(file)
                            xlwt_book = xlwt.Workbook()

                            if xlrd_book.nsheets > 1:
                                print(f"File '{file}' has multiple sheets")
                                modified = False

                                for sheet_name in xlrd_book.sheet_names():
                                    # If a sheet contains 'dataset' and doesnt contain 'dictionary' or 'note' then keep it even if it contains any keyword to delete
                                    if (
                                            'dataset' in sheet_name.lower() and 'dictionary' not in sheet_name.lower() and 'note' not in sheet_name.lower()) or (
                                            'LastDisplacement' in sheet_name.lower() and 'dictionary' not in sheet_name.lower() and 'note' not in sheet_name.lower()):
                                        print(
                                            f"Sheet '{sheet_name}' in '{file}' contains 'dataset' or 'LastDisplacement', keeping it")
                                        copy_worksheets(xlrd_book, xlwt_book, sheet_name)
                                    # If a sheet contains any of the keywords delete it
                                    elif any(regex.search(sheet_name) for regex in regexes_to_delete) or (
                                            'dataset' in sheet_name.lower() and 'dictionary' in sheet_name.lower()) or (
                                            'dataset' in sheet_name.lower() and 'note' in sheet_name.lower()):
                                        print(f"Removed sheet '{sheet_name}' from '{file}'")
                                        modified = True
                                    else:
                                        copy_worksheets(xlrd_book, xlwt_book, sheet_name)

                                if modified:
                                    xlwt_book.save(str(output_file))
                                    print(f"Saved modified .xls file: {output_file}")
                        except Exception as e:
                            print(f"Error processing .xls file {file}: {e}")
                        print('')
                    elif file_ext in ['.xlsx', '.xlsm']:
                        print(f"Processing file: {file}")

                        try:
                            wb = load_workbook(file)

                            if len(wb.sheetnames) > 1:
                                print(f"File '{file}' has multiple sheets")

                                for sheet_name in wb.sheetnames:
                                    if (
                                            'dataset' in sheet_name.lower() and 'dictionary' not in sheet_name.lower() and 'note' not in sheet_name.lower()) or (
                                            'LastDisplacement' in sheet_name.lower() and 'dictionary' not in sheet_name.lower() and 'note' not in sheet_name.lower()):
                                        print(
                                            f"Sheet '{sheet_name}' in '{file}' contains 'dataset' or 'LastDisplacement', keeping it")
                                    elif any(regex.search(sheet_name) for regex in regexes_to_delete) or (
                                            'dataset' in sheet_name.lower() and 'dictionary' in sheet_name.lower()) or (
                                            'dataset' in sheet_name.lower() and 'note' in sheet_name.lower()):
                                        wb.remove(wb[sheet_name])
                                        print(f"Removed sheet '{sheet_name}' from '{file}'")

                            wb.save(output_file)
                            print(f"Saved modified .xlsx file: {output_file}")

                        except Exception as e:
                            print(f"Error processing .xlsx file {file}: {e}")
                        print('')
                        time.sleep(1)

    # Step one:
    @staticmethod
    def move_folders_with_multisheet_excels(source_dir, target_dir):
        source_path = Path(source_dir)
        target_path = Path(target_dir)

        # Create target directory if it doesn't exist
        target_path.mkdir(parents=True, exist_ok=True)

        # Iterate over subfolders in the source directory
        for subdir in source_path.iterdir():
            if subdir.is_dir():
                move_folder = False

                # Iterate over files in the subfolder
                for file in subdir.glob('*.*'):
                    # Check if the file has an Excel extension
                    file_ext = file.suffix.lower()
                    if file_ext in ['.xlsx', '.xls', '.xlsm']:
                        # Check if the Excel file has more than one sheet
                        try:
                            if file_ext == '.xls':
                                engine = 'xlrd'
                            else:
                                engine = 'openpyxl'

                            xls = pd.read_excel(file, engine=engine, sheet_name=None)
                            if len(xls) > 1:
                                move_folder = True
                                print(f"Found multisheet Excel file: {file}")
                                break
                        except Exception as e:
                            print(f"Error reading Excel file {file}: {e}")

                # Move the folder if a multisheet Excel file was found
                if move_folder:
                    target_subdir = target_path / subdir.name
                    print(f"Moving folder '{subdir}' to '{target_subdir}' because it has multisheets")
                    shutil.move(str(subdir), str(target_subdir))
        time.sleep(5)

    # Step two:
    @staticmethod
    def seperate_excel_sheets_openpyxl(master_dir):
        # Iterate through the master folder
        for root, dirs, files in os.walk(master_dir):
            print(f"Processing folder for excel sheets seperation: {root}")

            # Iterate through files in the current folder
            for file in files:
                # Ignore temporary files
                if file.startswith('~$'):
                    continue

                # Check if the file is an Excel file
                if file.lower().endswith((".xlsx", ".xls", ".xlsm")):
                    print(f"Processing seperation of Excel file: {file}")
                    file_path = os.path.join(root, file)

                    # Load the Excel workbook
                    if file.lower().endswith(".xls"):
                        wb = xlrd.open_workbook(file_path)
                        sheet_names = wb.sheet_names()
                    else:
                        wb = load_workbook(file_path)
                        sheet_names = wb.sheetnames

                    # Check if the workbook has more than one sheet
                    if len(sheet_names) > 1:
                        print(f"Found {len(sheet_names)} sheets in the workbook")

                        # Iterate through sheets in the workbook
                        for sheet_name in sheet_names:
                            if file.lower().endswith(".xls"):
                                sheet = wb.sheet_by_name(sheet_name)
                            else:
                                sheet = wb[sheet_name]

                            # Create a new workbook for the current sheet
                            new_file_path = os.path.join(root, f"{os.path.splitext(file)[0]}_{sheet_name}.xlsx")
                            new_wb = xlsxwriter.Workbook(new_file_path)
                            new_sheet = new_wb.add_worksheet(name=sheet_name)

                            # Copy the data cell by cell
                            if file.lower().endswith(".xls"):
                                for row_idx in range(sheet.nrows):
                                    for col_idx in range(sheet.ncols):
                                        value = sheet.cell_value(row_idx, col_idx)
                                        new_sheet.write(row_idx, col_idx, value)
                            else:
                                for row_idx in range(1, sheet.max_row + 1):
                                    for col_idx in range(1, sheet.max_column + 1):
                                        value = sheet.cell(row_idx, col_idx).value
                                        new_sheet.write(row_idx - 1, col_idx - 1, value)

                            # Save the new workbook with the sheet name appended to the original file name
                            new_wb.close()
                            print(f"Saved sheet '{sheet_name}' as: {os.path.splitext(file)[0]}_{sheet_name}.xlsx")

                        # Remove the original file
                        os.remove(file_path)
                        print('Removed the original file that had all the sheets stacked')

    # or this one but better to use the one without pandas
    @staticmethod
    def separate_excel_sheets_pandas(folder):
        for root, _, files in os.walk(folder):
            for file in files:
                if file.lower().endswith(('.xls', '.xlsx', '.xlsm')):
                    file_path = os.path.join(root, file)
                    print(f"Processing file: {file_path}")

                    try:
                        # Read the Excel file
                        xls = pd.read_excel(file_path, sheet_name=None)

                        # Check if more than one sheet exists
                        if len(xls) > 1:
                            # Iterate through each sheet in the workbook
                            for sheet_name, sheet_data in xls.items():
                                # Print the current sheet being processed
                                print(f"Processing sheet: {sheet_name}")

                                # Save the sheet data into a new Excel file
                                new_file_path = os.path.splitext(file_path)[0] + f"_{sheet_name}" + \
                                                os.path.splitext(file_path)[1]
                                writer = pd.ExcelWriter(new_file_path, engine='openpyxl')
                                sheet_data.to_excel(writer, index=False, sheet_name=sheet_name)
                                writer.save()
                        else:
                            print(f"Only one sheet in the file: {file_path}. Skipping.")

                    except Exception as e:
                        print(f"Error processing file: {file_path}. Error: {e}")
        time.sleep(5)

    # Step three:
    @staticmethod
    def clean_excel_headers(master_dir):
        print(f"Cleaning excel headers of files in folder: {master_dir}")

        for root, _, files in os.walk(master_dir):
            for file in files:
                if file.startswith('~$'):
                    continue
                file_path = os.path.join(root, file)
                file_ext = os.path.splitext(file)[1].lower()

                if file_ext in ['.xlsx', '.xlsm', '.xls']:
                    print(f"Cleaning headers of file: {file_path}")

                    engine = 'openpyxl' if file_ext in ['.xlsx', '.xlsm'] else 'xlrd'

                    try:
                        # Read all sheets of the Excel file
                        excel_file = pd.read_excel(file_path, header=None, engine=engine, sheet_name=None)
                    except KeyError as e:
                        print(f"Error reading Excel file {file_path}: {e}")
                        continue

                    # Save the file as .xlsx
                    save_path = os.path.splitext(file_path)[0] + '.xlsx'

                    # Use the 'with' statement to automatically save and close the file
                    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                        for sheet_name, sheet_data in excel_file.items():
                            # Clean sheet headers
                            sheet_data = sheet_data.dropna(how='all').reset_index(drop=True)

                            max_non_empty_cells = 0
                            header_row = None

                            # Find the row with the most non-empty cells to use as the header
                            for index, row in sheet_data.iterrows():
                                non_empty_cells = sum(row.notna())

                                if non_empty_cells > max_non_empty_cells:
                                    max_non_empty_cells = non_empty_cells
                                    header_row = index

                            # Set the header row and remove it from the data
                            if header_row is not None:
                                sheet_data.columns = sheet_data.iloc[header_row]
                                sheet_data = sheet_data.iloc[header_row + 1:].reset_index(drop=True)

                                # Save the cleaned sheet with the original sheet name
                                sheet_data.to_excel(writer, sheet_name=sheet_name, index=False)
                            else:
                                print(f"No valid header row found in the file: {file_path}")

                    print(f"Cleaned headers and saved file: {save_path}")

                    # Delete the old .xls file if it was successfully saved as .xlsx
                    if file_ext == '.xls':
                        os.remove(file_path)
                        print(f"Deleted old .xls file: {file_path}")

    # Step four:
    @staticmethod
    def process_new_metadata_and_subfolders(master_dir):
        master_dir = Path(master_dir)
        for subfolder in master_dir.iterdir():
            if subfolder.is_dir():
                excel_files = list(subfolder.glob("*.xls*")) + list(subfolder.glob("*.XLS*"))
                metadata_file = next(subfolder.glob("Metadata_*.json"), None)

                if not metadata_file:
                    print(f"No metadata file found in {subfolder}")
                    continue

                with open(metadata_file, "r") as f:
                    metadata_content = json.load(f)

                moved_files = set()
                for excel_file in excel_files:
                    if excel_file.name.startswith('~$'):
                        continue
                    if excel_file in moved_files:
                        continue

                    # Read the sheet names from the Excel file
                    sheet_names = pd.read_excel(excel_file, sheet_name=None, engine=None).keys()

                    for sheet_name in sheet_names:
                        new_json_file = subfolder / f"{metadata_file.stem}_{sheet_name}.json"
                        metadata_content["table"] = f"{metadata_content['table']} {sheet_name}"

                        with open(new_json_file, "w", encoding='utf-8') as f:
                            json.dump(metadata_content, f, ensure_ascii=False, indent=4)

                        new_subfolder = subfolder.parent / f"{subfolder.name}_{sheet_name}".strip()
                        new_subfolder.mkdir(parents=True, exist_ok=True)
                        metadata_content["table"] = metadata_content["table"].replace(f" {sheet_name}", "")

                        try:
                            print(
                                f"Moving {excel_file} to {new_subfolder / excel_file.name} to its newly created folder with its metadata")
                            shutil.copy(excel_file, new_subfolder / excel_file.name)
                            shutil.move(new_json_file, new_subfolder / new_json_file.name)
                            moved_files.add(excel_file)
                        except FileNotFoundError as e:
                            print(f"Error moving file: {e}")

                os.remove(metadata_file)
                shutil.rmtree(subfolder)

    @staticmethod
    def run_all_cleaning_process(master_dir, master_dir_deleted_sheets, master_dir_multisheets):
        # Rename Files and Subfolders
        DataCleaner.rename_files_and_folders(master_dir)
        # Delete sheets with no data
        DataCleaner.delete_specific_sheets(master_dir, master_dir_deleted_sheets)
        # Move sheets with multiple data sheets to another folder
        DataCleaner.move_folders_with_multisheet_excels(master_dir_deleted_sheets, master_dir_multisheets)
        time.sleep(60)
        # Rename Files and Subfolders
        DataCleaner.rename_files_and_folders(master_dir_multisheets)
        # Seperate these sheets into excel files (with their sheet names)
        DataCleaner.seperate_excel_sheets_openpyxl(master_dir_multisheets)
        # Clean the headers of the 1 sheet files that are left in the main folder
        DataCleaner.clean_excel_headers(master_dir_deleted_sheets)
        # Clean the headers of the sheets that got seperated into files and moved to the multisheets folder
        DataCleaner.clean_excel_headers(master_dir_multisheets)
        # time.sleep(60)
        # Make new metadata and sub folders for these new files and delete the original folder (because it is now empty)
        master_dir_iterdir = Path(master_dir_multisheets)
        DataCleaner.process_new_metadata_and_subfolders(master_dir_iterdir)
